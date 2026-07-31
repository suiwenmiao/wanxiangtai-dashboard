#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Re-match category and sub-category for every historical row in the big table."""

from __future__ import annotations

import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from config import BASE_TABLE_PATH, BIG_TABLE_PATH, SHEET_NAME


def normalize_id(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").astype("Int64")


def main() -> None:
    print(f"匹配表: {BASE_TABLE_PATH}")
    print(f"大表: {BIG_TABLE_PATH}")
    base = pd.read_excel(BASE_TABLE_PATH)
    base_id_column = "主体ID" if "主体ID" in base.columns else "商品ID" if "商品ID" in base.columns else None
    required_base = {"品类", "细类"}
    missing_base = sorted(required_base - set(base.columns))
    if base_id_column is None:
        missing_base.append("主体ID/商品ID")
    if missing_base:
        raise SystemExit(f"匹配表缺少字段: {', '.join(missing_base)}")
    mapping = base[[base_id_column, "品类", "细类"]].rename(columns={base_id_column: "主体ID"}).copy()
    mapping["主体ID"] = normalize_id(mapping["主体ID"])
    mapping = mapping.dropna(subset=["主体ID"]).drop_duplicates("主体ID", keep="first")

    table = pd.read_excel(BIG_TABLE_PATH, sheet_name=SHEET_NAME)
    if "主体ID" not in table.columns:
        raise SystemExit("大表缺少主体ID字段，已停止，不会写入文件。")
    table["主体ID"] = normalize_id(table["主体ID"])
    before = table[["品类", "细类"]].copy() if {"品类", "细类"}.issubset(table.columns) else None
    table = table.drop(columns=[column for column in ["品类", "细类"] if column in table.columns])
    table = table.merge(mapping, on="主体ID", how="left", sort=False)
    table["品类"] = table["品类"].fillna("其他")
    table["细类"] = table["细类"].fillna("其他")

    backup = BIG_TABLE_PATH.with_name(f"{BIG_TABLE_PATH.stem}.匹配前备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}{BIG_TABLE_PATH.suffix}")
    shutil.copy2(BIG_TABLE_PATH, backup)

    workbook = load_workbook(BIG_TABLE_PATH)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"大表中不存在工作表: {SHEET_NAME}")
    worksheet = workbook[SHEET_NAME]
    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(list(table.columns))
    for row in table.itertuples(index=False, name=None):
        worksheet.append(list(row))
    workbook.save(BIG_TABLE_PATH)

    matched = int((table["品类"] != "其他").sum())
    changed = 0
    if before is not None and len(before) == len(table):
        changed = int(((before["品类"].fillna("其他").astype(str) != table["品类"].astype(str)) | (before["细类"].fillna("其他").astype(str) != table["细类"].astype(str))).sum())
    print(f"总行数: {len(table)}")
    print(f"匹配表主体数: {len(mapping)}")
    print(f"成功匹配: {matched}/{len(table)} ({matched / len(table) * 100:.1f}%)")
    print(f"品类/细类发生变化的行数: {changed}")
    print(f"备份文件: {backup}")
    print(f"已更新: {BIG_TABLE_PATH}")


if __name__ == "__main__":
    main()
